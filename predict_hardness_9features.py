import openpyxl
import math
import pandas as pd
import numpy as np

file_path = "calculate_features_and_hardness.xlsx"
df = pd.read_excel(file_path)
df.iloc[:, :15] = df.iloc[:, :15].fillna(0)
df.to_excel(file_path, index=False)

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [70,	209,	279,	130,	211,	200,	329,	116,	411,	105,	186,	128,	198,	78,	68]
    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        Young_modulus = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        average_deviation_of_Young_modulus = (
            abs(r_values[0] - Young_modulus) * row[0] + abs((r_values[1] - Young_modulus)) * row[1] +
            abs((r_values[2] - Young_modulus)) * row[2] + abs((r_values[3] - Young_modulus)) * row[3] +
            abs((r_values[4] - Young_modulus)) * row[4] + abs((r_values[5] - Young_modulus)) * row[5] +
            abs((r_values[6] - Young_modulus)) * row[6] + abs((r_values[7] - Young_modulus)) * row[7] +
            abs((r_values[8] - Young_modulus)) * row[8] + abs((r_values[9] - Young_modulus)) * row[9] +
            abs((r_values[10] - Young_modulus)) * row[10] + abs((r_values[11] - Young_modulus)) * row[11] +
            abs((r_values[12] - Young_modulus)) * row[12] + abs((r_values[13] - Young_modulus)) * row[13] +
            abs((r_values[14] - Young_modulus)) * row[14]
        ) * 0.01         
        sheet.cell(row=row_num, column=18).value = round(average_deviation_of_Young_modulus, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract values from the row
        # Calculate the result using the given formula
        VEC = (row[0]*3 + row[1]*9 + row[2]*6 + row[3]*11 + row[4]*8 + row[5]*10 + row[6]*6 + row[7]*4 + row[8]*6 + row[9]*5 +
                    row[10]*5 + row[11]*5 + row[12]*7 + row[13]*4 + row[14]*4)*0.01
        # Write the result to the 'T' column of the current row
        sheet.cell(row=row_num, column=19).value =round(VEC, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [0.2308, 0.3333, 0.25, 0.3793, 0.3077, 0.3571, 0.1429, 0.1818, 0.0808, 0.122, 0.0685, 0.2174, 0.28, 0.0556, 0.1]
    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        number_of_itinerant_electrons = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        average_deviation_of_number_of_itinerant_electrons = (
            abs(r_values[0] - number_of_itinerant_electrons) * row[0] + abs((r_values[1] - number_of_itinerant_electrons)) * row[1] +
            abs((r_values[2] - number_of_itinerant_electrons)) * row[2] + abs((r_values[3] - number_of_itinerant_electrons)) * row[3] +
            abs((r_values[4] - number_of_itinerant_electrons)) * row[4] + abs((r_values[5] - number_of_itinerant_electrons)) * row[5] +
            abs((r_values[6] - number_of_itinerant_electrons)) * row[6] + abs((r_values[7] - number_of_itinerant_electrons)) * row[7] +
            abs((r_values[8] - number_of_itinerant_electrons)) * row[8] + abs((r_values[9] - number_of_itinerant_electrons)) * row[9] +
            abs((r_values[10] - number_of_itinerant_electrons)) * row[10] + abs((r_values[11] - number_of_itinerant_electrons)) * row[11] +
            abs((r_values[12] - number_of_itinerant_electrons)) * row[12] + abs((r_values[13] - number_of_itinerant_electrons)) * row[13] +
            abs((r_values[14] - number_of_itinerant_electrons)) * row[14]
        ) * 0.01
        sheet.cell(row=row_num, column=20).value = round(number_of_itinerant_electrons, 4)
        sheet.cell(row=row_num, column=24).value = round(average_deviation_of_number_of_itinerant_electrons, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')


#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [143.17, 125.1, 124.91, 127.8, 124.12, 124.59, 136.26, 146.15, 136.7, 142.9, 143, 131.6, 135, 157.75, 160.25]
    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  
        atom_radius = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        average_deviation_of_atom_radius = (
            abs(r_values[0] - atom_radius) * row[0] + abs((r_values[1] - atom_radius)) * row[1] +
            abs((r_values[2] - atom_radius)) * row[2] + abs((r_values[3] - atom_radius)) * row[3] +
            abs((r_values[4] - atom_radius)) * row[4] + abs((r_values[5] - atom_radius)) * row[5] +
            abs((r_values[6] - atom_radius)) * row[6] + abs((r_values[7] - atom_radius)) * row[7] +
            abs((r_values[8] - atom_radius)) * row[8] + abs((r_values[9] - atom_radius)) * row[9] +
            abs((r_values[10] - atom_radius)) * row[10] + abs((r_values[11] - atom_radius)) * row[11] +
            abs((r_values[12] - atom_radius)) * row[12] + abs((r_values[13] - atom_radius)) * row[13] +
            abs((r_values[14] - atom_radius)) * row[14]
        ) * 0.01
        sheet.cell(row=row_num, column=21).value = round(average_deviation_of_atom_radius, 4)
        sheet.cell(row=row_num, column=25).value = round(atom_radius, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#####################################################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [143.17, 125.1, 124.91, 127.8, 124.12, 124.59, 136.26, 146.15, 136.7, 142.9, 143, 131.6, 135, 157.75, 160.25]   
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        r_average = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        daierta1 = (
            (1 - r_values[0] / r_average) ** 2 * row[0] + (1 - (r_values[1] / r_average))**2 * row[1] +
            (1 - (r_values[2] / r_average))**2 * row[2] + (1 - (r_values[3] / r_average))**2 * row[3] +
            (1 - (r_values[4] / r_average))**2 * row[4] + (1 - (r_values[5] / r_average))**2 * row[5] +
            (1 - (r_values[6] / r_average))**2 * row[6] + (1 - (r_values[7] / r_average))**2 * row[7] +
            (1 - (r_values[8] / r_average))**2 * row[8] + (1 - (r_values[9] / r_average))**2 * row[9] +
            (1 - (r_values[10] / r_average))**2 * row[10] + (1 - (r_values[11] / r_average))**2 * row[11] +
            (1 - (r_values[12] / r_average))**2 * row[12] + (1 - (r_values[13] / r_average))**2 * row[13] +
            (1 - (r_values[14] / r_average))**2 * row[14]
        ) * 0.01
        daierta=daierta1**0.5
        # Write the result to the 'T' column of the current row
        sheet.cell(row=row_num, column=22).value = round(daierta, 4)   
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

############### Define the formula
def custom_formula(row):
    return 0.0004 * (
        row[0]*row[1]*(-17.11) + row[0]*row[2]*(-9.321) + row[0]*row[3]*(-7.149) + row[0]*row[4]*(-10.4) +
        row[0]*row[5]*(-20.67) + row[0]*row[6]*(-4.981) + row[0]*row[7]*(-29.12) + row[0]*row[8]*(-1.939) +
        row[0]*row[9]*(-17.89) + row[0]*row[10]*(-18.84) + row[0]*row[11]*(-15.71) + row[0]*row[12]*(-18.02) +
        row[0]*row[13]*(-36.44) + row[0]*row[14]*(-41.1) + row[1]*row[2]*(-4.383) + row[1]*row[3]*(6.321) +
        row[1]*row[4]*(-0.559) + row[1]*row[5]*(-0.2176) + row[1]*row[6]*(-4.6) + row[1]*row[7]*(-25.93) +
        row[1]*row[8]*(-1.317) + row[1]*row[9]*(-22.41) + row[1]*row[10]*(-21.85) + row[1]*row[11]*(-13.36) +
        row[1]*row[12]*(-5.077) + row[1]*row[13]*(-30.5) + row[1]*row[14]*(-35.14) + row[2]*row[3]*(12.36) +
        row[2]*row[4]*(-1.447) + row[2]*row[5]*(-6.546) + row[2]*row[6]*(0.3607) + row[2]*row[7]*(-6.945) +
        row[2]*row[8]*(0.9148) + row[2]*row[9]*(-6.675) + row[2]*row[10]*(-6.235) + row[2]*row[11]*(-1.905) +
        row[2]*row[12]*(2.102) + row[2]*row[13]*(-8.251) + row[2]*row[14]*(-10.66) + row[3]*row[4]*(12.82) +
        row[3]*row[5]*(3.481) + row[3]*row[6]*(17.55) + row[3]*row[7]*(-8.279) + row[3]*row[8]*(21.22) +
        row[3]*row[9]*(2.403) + row[3]*row[10]*(1.73) + row[3]*row[11]*(4.812) + row[3]*row[12]*(3.712) +
        row[3]*row[13]*(-14.97) + row[3]*row[14]*(-19.96) + row[4]*row[5]*(-1.527) + row[4]*row[6]*(-1.882) +
        row[4]*row[7]*(-15.55) + row[4]*row[8]*(-0.0452) + row[4]*row[9]*(-14.48) + row[4]*row[10]*(-13.84) +
        row[4]*row[11]*(-6.901) + row[4]*row[12]*(0.2259) + row[4]*row[13]*(-18.19) + row[4]*row[14]*(-21.69) +
        row[5]*row[6]*(-6.868) + row[5]*row[7]*(-31.58) + row[5]*row[8]*(-2.938) + row[5]*row[9]*(-27.21) +
        row[5]*row[10]*(-26.64) + row[5]*row[11]*(-17.19) + row[5]*row[12]*(-8.009) + row[5]*row[13]*(-36.97) +
        row[5]*row[14]*(-42.14) + row[6]*row[7]*(-3.531) + row[6]*row[8]*(-0.21) + row[6]*row[9]*(-5.495) +
        row[6]*row[10]*(-4.786) + row[6]*row[11]*(-0.0287) + row[6]*row[12]*(4.675) + row[6]*row[13]*(-3.69) +
        row[6]*row[14]*(-5.789) + row[7]*row[8]*(-5.583) + row[7]*row[9]*(1.96) + row[7]*row[10]*(1.374) +
        row[7]*row[11]*(-1.595) + row[7]*row[12]*(-7.611) + row[7]*row[13]*(0.1584) + row[7]*row[14]*(-0.2271) +
        row[8]*row[9]*(-8.1) + row[8]*row[10]*(-7.161) + row[8]*row[11]*(-7.865) + row[8]*row[12]*(5.981) +
        row[8]*row[13]*(-5.961) + row[8]*row[14]*(-8.476) + row[9]*row[10]*(0.0265) + row[9]*row[11]*(-0.9903) +
        row[9]*row[12]*(-3.467) + row[9]*row[13]*(3.754) + row[9]*row[14]*(3.732) + row[10]*row[11]*(-0.964) +
        row[10]*row[12]*(-3.564) + row[10]*row[13]*(2.792) + row[10]*row[14]*(2.575) + row[11]*row[12]*(-0.697) +
        row[11]*row[13]*(-1.985) + row[11]*row[14]*(-3.374) + row[12]*row[13]*(-10.61) + row[12]*row[14]*(-13.59) +
        row[13]*row[14]*(-0.2139)
    )
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        enthalpy = round(custom_formula(row), 4)
        # Write the total to the 'R' column of the current row
        sheet.cell(row=row_num, column=23).value = enthalpy
   # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [1.79, 7.88, 8.62, 4.02, 8.85, 7.26, 9.17, 5.23, 13.38, 7.01, 8.22, 7.66, 6.64, 5.26, 3.94]    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        specific_heat_capacity = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        average_deviation_of_specific_heat_capacity = (
            abs(r_values[0] - specific_heat_capacity) * row[0] + abs((r_values[1] - specific_heat_capacity)) * row[1] +
            abs((r_values[2] - specific_heat_capacity)) * row[2] + abs((r_values[3] - specific_heat_capacity)) * row[3] +
            abs((r_values[4] - specific_heat_capacity)) * row[4] + abs((r_values[5] - specific_heat_capacity)) * row[5] +
            abs((r_values[6] - specific_heat_capacity)) * row[6] + abs((r_values[7] - specific_heat_capacity)) * row[7] +
            abs((r_values[8] - specific_heat_capacity)) * row[8] + abs((r_values[9] - specific_heat_capacity)) * row[9] +
            abs((r_values[10] - specific_heat_capacity)) * row[10] + abs((r_values[11] - specific_heat_capacity)) * row[11] +
            abs((r_values[12] - specific_heat_capacity)) * row[12] + abs((r_values[13] - specific_heat_capacity)) * row[13] +
            abs((r_values[14] - specific_heat_capacity)) * row[14]
        ) * 0.01
        sheet.cell(row=row_num, column=26).value = round(average_deviation_of_specific_heat_capacity, 4)
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [25, 82.6, 110, 46, 77.5, 76, 120, 43, 156, 37.5, 69, 46.4, 76.4, 30, 35.3]    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        Shear_Modulus = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        # Write the result to the 'T' column of the current row
        sheet.cell(row=row_num, column=27).value = round(Shear_Modulus, 4)
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')


#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [933, 1768, 2180, 1358, 1811, 1728, 2896, 1941, 3695, 2750, 3290, 2183, 1519, 2506, 2128]   
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Count non-zero values in the first 15 columns
        non_zero_count = sum(1 for value in row[:15] if value != 0)    	  	    
        Tm = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        average_deviation_of_Tm = (
            abs(r_values[0] - Tm) * row[0] + abs((r_values[1] - Tm)) * row[1] +
            abs((r_values[2] - Tm)) * row[2] + abs((r_values[3] - Tm)) * row[3] +
            abs((r_values[4] - Tm)) * row[4] + abs((r_values[5] - Tm)) * row[5] +
            abs((r_values[6] - Tm)) * row[6] + abs((r_values[7] - Tm)) * row[7] +
            abs((r_values[8] - Tm)) * row[8] + abs((r_values[9] - Tm)) * row[9] +
            abs((r_values[10] - Tm)) * row[10] + abs((r_values[11] - Tm)) * row[11] +
            abs((r_values[12] - Tm)) * row[12] + abs((r_values[13] - Tm)) * row[13] +
            abs((r_values[14] - Tm)) * row[14]
        ) * 0.01
        sheet.cell(row=row_num, column=28).value = round(average_deviation_of_Tm, 4)
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')


##################################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=15, values_only=True), start=2):
        entropy = 0
        for cell_value in row:
            # Stop calculating if cell is empty or 0
            if cell_value is None or cell_value == 0:
                continue
            entropy += -round(8.314 * (cell_value / 100) * math.log(cell_value / 100), 4)
            
        # Write the total to the 'S' column of the current row
        sheet.cell(row=row_num, column=29).value = round(entropy, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract values from the row
        # Calculate the result using the given formula
        oumeiga = (((row[0]*933 + row[1]*1768 + row[2]*2180 + row[3]*1358 + row[4]*1811 + row[5]*1728 + row[6]*2896 + row[7]*1941 + row[8]*3695 + row[9]*2750 +
                    row[10]*3290 + row[11]*2183 + row[12]*1519 + row[13]*2506 + row[14]*2128) * 0.01) * row[28] / abs(row[22])) * 0.001
        # Write the result to the 'T' column of the current row
        sheet.cell(row=row_num, column=30).value =round(oumeiga, 4)
    # Save the workbook
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [1.61, 1.88, 1.66, 1.9, 1.83, 1.91, 2.16, 1.54, 2.36, 1.6, 1.5, 1.63, 1.55, 1.3, 1.33]
    
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Count non-zero values in the first 15 columns
        non_zero_count = sum(1 for value in row[:15] if value != 0)      
        Pauling_Electronegativity = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        sheet.cell(row=row_num, column=31).value = round(Pauling_Electronegativity, 4)
    # Save the workbook
    wb.save(filename)

# Replace 'The calculation of features.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

#############################################
def calculate_feature(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    r_values = [3,	2,	1,	1,	2,	2,	1,	2,	2,	1,	2,	2,	2,	2,	2]
    # Iterate through each row starting from the second row
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):       
        e_a = (
            row[0] * r_values[0] + row[1] * r_values[1] + row[2] * r_values[2] +
            row[3] * r_values[3] + row[4] * r_values[4] + row[5] * r_values[5] +
            row[6] * r_values[6] + row[7] * r_values[7] + row[8] * r_values[8] +
            row[9] * r_values[9] + row[10] * r_values[10] + row[11] * r_values[11] +
            row[12] * r_values[12] + row[13] * r_values[13] + row[14] * r_values[14]
        ) * 0.01
        sheet.cell(row=row_num, column=32).value = round(e_a, 4)
        
    wb.save(filename)
# Replace 'calculate_features_and_hardness.xlsx' with the actual filename
calculate_feature('calculate_features_and_hardness.xlsx')

###############predict_hardness
import pandas as pd  
import joblib  

test_data = pd.read_excel('calculate_features_and_hardness.xlsx')    
input_data = test_data.iloc[:, 17:26]   
model = joblib.load('XGBoost_6.pkl')  

predictions = model.predict(input_data)  

test_data.iloc[:, 15] = predictions  
test_data.to_excel('calculate_features_and_hardness.xlsx', index=False)